#!/usr/bin/python
# -*- coding: utf-8 -*-

import fire
from zipfile import ZipFile, is_zipfile
import StringIO
import fileinput
import sys
import gzip
import binascii
import pandas
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import OrderedDict
import arrow
import os, fnmatch

# define battery drop level
BATTERY_DROP_WATER_LEVEL = 6


NETLOG_TITLE = 'local_time  timestamp	clientAddress	logType	formatVersion	clientBytesIn	' \
                'clientBytesOut	serverBytesIn	serverBytesOut	cacheBytesIn	cacheBytesOut	' \
                'host	application	applicationStatus	operation	protocolStack	networkProtocolStack	' \
                'networkInterface	responseDelay	responseDuration	requestId	subscriptionId	' \
                'statusCode	errorCode	contentType	headerLength	contentLength	responseHash	' \
                'analysisString	analysis	optimization	protocolDetection	destinationIp	' \
                'destinationPort	redirectedToIp	redirectedToPort	sequenceNumber	requestDelay	' \
                'radioAwarenessStatus	originatorId	csmCreationTime	clientSrcPort	cspSrcPort'


def is_gz_file(filepath):
    with open(filepath, 'rb') as test_f:
        return binascii.hexlify(test_f.read(2)) == b'1f8b'

def open_file_input_string(input_file):
    if is_gz_file(input_file):
        with gzip.open(input_file, 'rb') as f:
            file_content = f.read()
        file_object = StringIO.StringIO(file_content)
    else:
        file_object = fileinput.input(input_file)

    return file_object




def parse_crcs_from_file(f):
    global df_all,df_power,df_netlog,df_system,df_backlight,df_deviceinfo,df_memory,df_memory_power_fast
    global crcs_start_time,crcs_end_time,user_id,orig_file,total_crcs_number
    global basic_report,device_time_zone


    df_all = pandas.DataFrame()
    df_system = pandas.DataFrame()
    df_netlog = pandas.DataFrame()
    df_power = pandas.DataFrame()
    df_memory = pandas.DataFrame()
    df_memory_power_fast = pandas.DataFrame()
    basic_report = OrderedDict()

    file_object = open_file_input_string(f)
    df_all = pandas.read_csv(file_object,header=None,names = list(range(0,50)))
    print "read_csv success,record number:" + str(df_all.shape[0])

    if df_all.empty:
        return

    df_all.dropna(axis=1,how="all",inplace=True)

    # save to excel
    ws_all = wb.create_sheet("all")



    total_crcs_number = df_all.shape[0]
    df_all[0] = pandas.to_datetime(df_all.iloc[:,0],format="%Y-%m-%d %H:%M:%S",errors='coerce')
    crcs_start_time = df_all.iloc[0,0]
    crcs_end_time = df_all.iloc[-1,0]
    orig_file = f
    user_id = df_all.iloc[0,1]


    # battery data frame
    df_power = df_all[df_all.iloc[:, 4] == 'battery'].copy()
    if df_power.empty:
        return
    df_power[5] = pandas.to_numeric(df_power[5])
    df_power[6] = pandas.to_numeric(df_power[6])
    df_power[7] = pandas.to_numeric(df_power[7])/1000
    df_power[8] = pandas.to_numeric(df_power[8])
    df_power[9] = pandas.to_numeric(df_power[9])

    # get the fast power records.
    df_netlog = df_all[df_all[2] == 'netlog'].copy()

    # process system log
    df_system = df_all[df_all[2] == 'system']
    df_backlight = df_system[df_system[4] == 'backlight'].copy()
    df_backlight[5] = pandas.to_numeric(df_backlight[5])
    df_backlight[10] = pandas.to_numeric(df_backlight[10])

    df_deviceinfo = df_system[df_system[4] == 'dev_info'].copy()
    df_memory = df_system[df_system[4] == 'memory'].copy()

    calc_device_info()

    ws_all.append(NETLOG_TITLE.split())

    for r in dataframe_to_rows(df_all, index=False, header=False):
        local_time = arrow.get(r[0]).to(device_time_zone)
        r.insert(0, local_time.naive)
        ws_all.append(r)
    pass


    process_cpu_logs()
    process_memory_logs()
    process_power_fast_logs()
    process_service_logs()
    generate_basic_battery_report()



def calc_screen_battery_usage():
    global avg_battery_drop_speed_screen_on,avg_battery_drop_speed_screen_off
    avg_battery_drop_speed_screen_on = 0
    avg_battery_drop_speed_screen_off = 0
    df_power_backlight = df_all[df_all[4].isin(['battery','backlight'])].copy()
    df_power_backlight = df_power_backlight.reset_index()
    df_power_screen_on = pandas.DataFrame()
    df_power_screen_off = pandas.DataFrame()
    df_power_temp = pandas.DataFrame()
    pre_screen_state = -1
    now_screen_events = -1
    for index,row in df_power_backlight.iterrows():
        if row[2] == 'power':
            df_power_temp = df_power_temp.append(row)
        elif row[2] == 'system':
            if int(row[5]) == 100:
                now_screen_events = 1
            elif int(row[5]) == 0:
                now_screen_events = 0

            if now_screen_events == pre_screen_state:
                continue

            if now_screen_events:
                df_power_screen_off = df_power_screen_off.append(df_power_temp)
            else:
                df_power_screen_on = df_power_screen_on.append(df_power_temp)

            pre_screen_state = now_screen_events
            df_power_temp.drop(df_power_temp.index, inplace=True)



            pass
        # end if
        pass
    # end for
    if df_power_screen_on.shape[0] != 0:
        df_power_screen_on[7] = pandas.to_numeric(df_power_screen_on[7]) / 1000
        avg_battery_drop_speed_screen_on = df_power_screen_on[7].mean()\

    if df_power_screen_off.shape[0] != 0:
        df_power_screen_off[7] = pandas.to_numeric(df_power_screen_off[7]) / 1000
        avg_battery_drop_speed_screen_off = df_power_screen_off[7].mean()

    pass


def calc_device_info():
    global device_mode,device_version,device_time_zone
    global df_deviceinfo
    device_mode = "unknown"
    device_version = "unknown"
    device_time_zone = "UTC"
    df_device = df_deviceinfo[df_deviceinfo[5] == '0']
    df_time_zone = df_deviceinfo[df_deviceinfo[5] == "time_zone"]
    if not df_device.empty:
        device_mode = df_device.iloc[0,:][6]
        device_version = df_device.iloc[0,:][7]
        pass

    if not df_time_zone.empty:
        device_time_zone = df_time_zone.iloc[0,:][7]
        pass



def generate_basic_battery_report():
    global basic_report,df_memory,df_memory_power_fast
    global device_mode, device_version, device_time_zone

    avg_battery_drop_speed = df_power[7].mean()
    calc_screen_battery_usage()
    basic_report['user'] = user_id
    basic_report['total_crcs_number'] = total_crcs_number
    basic_report['ave_battery_speed'] \
        = '{}s {}%/h'.format(int(avg_battery_drop_speed),round(3600/avg_battery_drop_speed,2))
    basic_report['screen_on_power_speed'] = "NA"
    basic_report['screen_off_power_speed'] = "NA"

    if avg_battery_drop_speed_screen_on != 0:
        basic_report['screen_on_power_speed'] \
            = '{}s {}%/h'.format(int(avg_battery_drop_speed_screen_on),round(3600/avg_battery_drop_speed_screen_on,2))


    if avg_battery_drop_speed_screen_off:
        basic_report['screen_off_power_speed'] \
            = '{}s {}%/h'.format(int(avg_battery_drop_speed_screen_off),round(3600/avg_battery_drop_speed_screen_off,2))

    basic_report['device_mode'] = device_mode
    basic_report['device_version'] = device_version
    basic_report['time_zone'] = device_time_zone
    basic_report['time_period'] = str(crcs_end_time - crcs_start_time)
    basic_report['start_time_utc'] = crcs_start_time
    basic_report['end_time_utc'] = crcs_end_time
    basic_report['memory'] = generate_memory_report(df_memory)
    basic_report['memory_power_fast'] = generate_memory_report(df_memory_power_fast)
    basic_report['source_file'] = orig_file



    # if not title_exist:
    #     ws_basic.append(result.keys())
    #     title_exist = True
    # ws_basic.append(result.values())

    for k,v in basic_report.items():
        ws_basic.append([k,v])

    pass

def process_cpu_logs():
    df_cpu = df_system[df_system[4] == 'cpu'].copy()
    if df_cpu.empty:
        return
    ws_cpu = wb.create_sheet("cpu")

    for r in dataframe_to_rows(df_cpu, index=False, header=False):
        if "[process]" in r[6]:
            ws_cpu.append(r)

    df_cpu_filter = df_cpu[df_cpu[6].str.contains("process")].copy()

    process_list = df_cpu_filter[5].unique()

    for process in process_list:
        df_process_cpu = df_cpu_filter[df_cpu_filter[5] == process]
        s_process_cpu = df_process_cpu[7].str.split("/")

        pass

    pass

def process_power_fast_logs():
    global df_memory_power_fast
    df_power_fast = df_power[df_power[7] <= BATTERY_DROP_WATER_LEVEL*60].copy()
    if df_power_fast.empty:
        return
    df_power_fast['start_time'] = df_power_fast[0] - pandas.to_timedelta(df_power_fast[7],'s')
    df_power_fast['end_time'] = df_power_fast[0]

    ws_power_fast = wb.create_sheet("battery_drop_fast")
    ws_power_netlog = wb.create_sheet("netlog_battery_drop_fast")

    df_netlog_power_fast = pandas.DataFrame()

    for index, value in df_power_fast.iterrows():
        start_time = value['start_time']
        end_time = value['end_time']
        df_netlog_temp = df_netlog[(df_netlog[0] >= start_time) & (df_netlog[0] <= end_time)]
        df_netlog_power_fast = df_netlog_power_fast.append(df_netlog_temp)
        if not df_memory.empty:
            df_memory_power_fast = df_memory_power_fast.append(
                df_memory[(df_memory[0] >= start_time) & (df_memory[0] <= end_time)])

    df_power_fast = df_power_fast.append(df_backlight)
    df_power_fast = df_power_fast.append(df_memory_power_fast)
    df_power_fast = df_power_fast.sort_values(df_power_fast.columns[0])

    cols = df_netlog_power_fast.columns[df_netlog_power_fast.dtypes.eq('object')]
    df_netlog_power_fast[cols] = df_netlog_power_fast[cols].apply(pandas.to_numeric,errors="ignore")

    for r in dataframe_to_rows(df_power_fast, index=False, header=False):
        ws_power_fast.append(r)
    pass

    for r in dataframe_to_rows(df_netlog_power_fast, index=False, header=False):
        ws_power_netlog.append(r)




def process_service_logs():
    df_service = df_all[(df_all[2] == 'service') & (df_all[4] == 'service')].copy()
    if df_service.empty:
        return

    ws_service = wb.create_sheet("service_log")

    for r in dataframe_to_rows(df_service,index=False,header=False):
        ws_service.append(r)

    pass


def generate_memory_report(_df_memory):
    str_memory_info = ""

    process_list = _df_memory[6].unique()

    for process in process_list:
        s_memory_info = _df_memory[_df_memory[6] == process][11]
        df_process_memory = s_memory_info.str.split(pat=":", expand=True)
        df_process_memory[1] = pandas.to_numeric(df_process_memory[1])
        mean = round(df_process_memory[1].mean(), 2)
        max = round(df_process_memory[1].max(), 2)
        str_memory_info += "[{}]:mean - {}, max - {} \n".format(process, mean, max)
        pass


    return str_memory_info

def process_memory_logs():
    global df_memory


    if df_memory.empty:
        return

    df_memory = df_memory[df_memory[5] == "process"].copy()
    ws_memory = wb.create_sheet("memory")

    for r in dataframe_to_rows(df_memory, index=False, header=False):
        ws_memory.append(r)
    pass






def find_files(directory, pattern):
    for root, dirs, files in os.walk(directory):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                yield filename



def parser_crcs_from_dir(f):
    for filename in find_files(f, '*.*'):
        print 'Found file:', filename
        parse_crcs_from_file(filename)


        #gc.collect()
    pass

class CRCSAnaLyser(object):
    def parse(self,f):
        if os.path.isdir(f):
            parser_crcs_from_dir(f)
        else:
            parse_crcs_from_file(f)
        pass

    def test_parse_crcs_file(self):
        parse_crcs_from_file('data/Crcs_#7515.gz')

    def test_parse_crcs_folder(self):
        parser_crcs_from_dir('data/crcs_test_folder')
        pass



def main():
    global wb,ws_basic
    wb = Workbook()
    ws_basic = wb.active
    ws_basic.title = "basic battery info"
    fire.Fire(CRCSAnaLyser)
    wb.save("battery_report_" + arrow.now().format("YYYY_MM_DD") + ".xlsx")


def on_file_uploaded(file,output_path):
    global wb, ws_basic
    wb = Workbook()
    ws_basic = wb.active
    ws_basic.title = "basic battery info"
    parse_crcs_from_file(file)
    filename = user_id + ".xlsx"
    wb.save(output_path + "/" + filename)
    return filename

if __name__ == "__main__":
    main()