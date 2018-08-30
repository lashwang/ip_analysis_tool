#!/usr/bin/python
# -*- coding: utf-8 -*-
import fire
import re
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'logcat_output.xlsx'
ws_csm = wb.active
ws_csm.title = "csm"
ws_netlog = wb.create_sheet(title="netlog")
ws_logcat_all = wb.create_sheet(title="logcat_all")
ws_logcat_java = wb.create_sheet(title="logcat_java")
ws_logcat_crash = wb.create_sheet(title="logcat_crash")
ws_thread_pool = wb.create_sheet(title="thread_pool")

ws_thread_pool.append("csm_id,task_name,task_id,time,pending_number".split(","))


csm_list = []

NET_LOG_HEADERS_V15 = ['timestamp','logType', 'formatVersion', \
                       'clientBytesIn', 'clientBytesOut', 'serverBytesIn', 'serverBytesOut', \
                       'cacheBytesIn', 'cacheBytesOut', 'host', 'application', 'applicationStatus', \
                       'operation', 'protocolStack', 'networkProtocolStack', 'networkInterface', \
 \
                       'responseDelay', 'responseDuration', 'requestId', 'subscriptionId', 'statusCode', \
                       'errorCode', 'contentType', 'headerLength', 'contentLength', 'responseHash', \
 \
                       'analysisString', 'analysis', 'optimization', 'protocolDetection', 'destinationIp',
                       'destinationPort', \
                       'redirectedToIp', 'redirectedToPort', 'sequenceNumber', 'requestDelay', 'radioAwarenessStatus', \
                       'originatorId', 'csmCreationTime', 'clientSrcPort', 'cspSrcPort'
                       ]


CSM_LOG_HEADER = ['date', 'time','pid', 'tid', 'csm', 'filename', 'errcode', 'log']

LOGCAT_HEADER = ['date', 'time', 'module','pid','tid','filename', 'errcode', 'log']

LOGCAT_HEADER_JAVA = ['date', 'time', 'pid','tid','classname','log']

USELESS_CLASS_JAVA_LOGS = [
    "ForegroundAppMonitor",
    "Z7AlarmManagerImpl",
    "UtilService",
    "LogHandler",
]

USELESS_FILE_CPP_LOGS = [
    "decoder_ifc.c",
    "Decoder.hpp",
    "encoder_ifc.c",
    "codec_plain_reader.c",
    "codec_processor_reader.c",
    "codec_plain_writer.c",
    "Protocol.cpp",
    "Protocol.hpp"
]

old_logcat_format = False


def parser_line(line):

    global old_logcat_format
    match_for_native = False
    match_for_csm = False


    reg_str = r"\s+F\s+\w+\s*\:\s*(.*)"
    matchObj = re.search(reg_str, line)
    if matchObj:
        log = matchObj.group(1)
        ws_logcat_crash.append([log])
        return




    if "[Native]" not in line and "[JAVA]" not in line:
        return


    line = line.strip()
    pid = -1
    tid = -1

    # get common header
    if not old_logcat_format:
        reg_str = r"^(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+"
        matchObj = re.search(reg_str, line)
        if matchObj:
            pid = matchObj.group(3)
            tid = matchObj.group(4)
            if not pid.isdigit():
                old_logcat_format = True
        else:
            old_logcat_format = True

    if old_logcat_format:
        reg_str = r"^\S+\s+\S+\s+\S\/\S+\(\s*(\S+)\):.*"
        matchObj = re.search(reg_str, line)
        if matchObj:
            pid = matchObj.group(1)
        pass

    # get native string.
    reg_str \
        = r"\[Native\]\S+\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+\S+\s+\[(\S+):(\S+)\]\s+\((\S+)\)\s+-\s+CSM\s+\[(\S+)\]\s+(.*)"
    matchObj = re.search(reg_str,line)
    if matchObj:
        index = 1
        date = matchObj.group(index)
        index += 1
        time = matchObj.group(index)
        index += 1
        time_zone = matchObj.group(index)
        index += 1
        tid = matchObj.group(index)
        index += 1
        filename = matchObj.group(index)
        index += 1
        flie_line = matchObj.group(index)
        index += 1
        errcode = matchObj.group(index)
        index += 1
        csm = matchObj.group(index)
        index += 1
        log = matchObj.group(index)
        ws_csm.append([date, time, pid,tid, csm, filename + ":" + flie_line, errcode, log])
        match_for_csm = True

        # if csm not in csm_list:
        #     csm_list.append(csm)



        # if csm not in csm_dict:
        #     csm_dict[csm] = wb.create_sheet(title="" + csm)
        #
        # csm_dict[csm].append([date, time, tid, csm, filename, flie_line, errcode, log])

    reg_str = r"NetLog\s+\(.*\):\s+(.*)"
    matchObj = re.search(reg_str, line)
    if matchObj:
        netlog_str = matchObj.group(1)
        ws_netlog.append(netlog_str.split(','))

    reg_str = r"\[Native\](\S+):\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s\[\S+\]\s+\[(\S+):(\S+)\]\s\((\S+)\)\s+-\s+(.*)"
    matchObj = re.search(reg_str, line)
    if matchObj:
        index = 1
        module = matchObj.group(index)
        index += 1
        date = matchObj.group(index)
        index += 1
        time = matchObj.group(index)
        index += 1
        time_zone = matchObj.group(index)
        index += 1
        tid = matchObj.group(index)
        index += 1
        filename = matchObj.group(index)
        index += 1
        flie_line = matchObj.group(index)
        index += 1
        errcode = matchObj.group(index)
        index += 1
        log = matchObj.group(index)
        if filename not in USELESS_FILE_CPP_LOGS:
            ws_logcat_all.append([date, time,module,pid, tid, filename + ":" + flie_line, errcode, log])
        match_for_native = True


    if match_for_native and not match_for_csm:
        reg_str = r"(.*)\[([0-9a-fA-F]{8})\](.*)"
        matchObj = re.search(reg_str, log)
        if matchObj:
            csm = matchObj.group(2)
            ws_csm.append([date, time, pid, tid, csm, filename + ":" + flie_line, errcode, log])

    reg_str \
        = "\[threadpool.cpp:(\d+)\](.+)Executed task\s+(.+)\((\d+)\)\,\s+.*CSM\s+\[(\S+)\].*take\s+(\d+)ms,pending task:(\d+)"
    matchObj = re.search(reg_str, line)
    if matchObj:
        task_name = matchObj.group(3)
        task_code = matchObj.group(4)
        csm_id = matchObj.group(5)
        time = matchObj.group(6)
        pending_task_number = matchObj.group(7)
        ws_thread_pool.append([csm_id,task_name,task_code,time,pending_task_number])
        #print line

    # get java related logs.
    if not match_for_native:
        reg_str = r"\[JAVA\](\S+):\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+\S+\s+(.*)"
        matchObj = re.search(reg_str, line)
        if matchObj:
            index = 1
            classname = matchObj.group(index)
            index += 1
            date = matchObj.group(index)
            index += 1
            time = matchObj.group(index)
            index += 1
            time_zone = matchObj.group(index)
            index += 1
            tid = matchObj.group(index)
            index += 1
            log = matchObj.group(index)
            index += 1
            if classname not in USELESS_CLASS_JAVA_LOGS:
                ws_logcat_java.append([date,time,pid,tid,classname,log])


    pass


def parse_file(fname):
    dest_filename = fname + ".xlsx";

    with open(fname) as f:
        content = f.readlines()


    ws_csm.append(CSM_LOG_HEADER)
    ws_netlog.append(NET_LOG_HEADERS_V15)
    ws_logcat_all.append(LOGCAT_HEADER)
    ws_logcat_java.append(LOGCAT_HEADER_JAVA)
    for line in content:
        try:
            parser_line(line)
        except Exception,e:
            print "error:" + e

    wb.save(filename=dest_filename)

    pass


def main():
    fire.Fire(parse_file)
    pass


if __name__ == "__main__":
    main()
